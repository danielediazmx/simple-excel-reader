import pathlib

import openpyxl
import sys
import requests
import json


class excelTransformer:
    def __init__(self, filename, g_key):
        self.filename = filename
        self.g_key = g_key
        self.wb_obj = openpyxl.load_workbook(filename)
        self.sheet_obj = self.wb_obj.active
        self.max_row = self.sheet_obj.max_row

        self.already_writted = []

    def write_address_in_file(self, address, row):
        mycell = self.sheet_obj['F' + str(row)]
        mycell.value = address

    def get_address_from_data(self, json_data, row):
        for data in json_data:
            types = data['types']

            if row in self.already_writted:
                continue

            if 'political' in types:
                self.write_address_in_file(data['formatted_address'], row)
                self.already_writted.append(row)

    def get_info_from_gapi(self, latitud, longitud, row):
        api_result = requests.get(
            "https://maps.google.com/maps/api/geocode/json?latlng=" + latitud + "," + longitud + "&key=" + self.g_key,
            stream=True).raw.data
        json_object = json.loads(api_result)

        self.get_address_from_data(json_object["results"], row)

    def isMultiple(self, num, check_with):
        return num % check_with == 0

    def read_excel(self):
        for i in range(2, self.max_row + 1):
            latitud = self.sheet_obj.cell(row=i, column=1)
            longitud = self.sheet_obj.cell(row=i, column=2)
            address = self.sheet_obj.cell(row=i, column=6)

            if longitud.value and latitud.value and not address.value:
                latitud = str(latitud.value)
                longitud = str(longitud.value)

                if self.isMultiple(i, 500):
                    self.wb_obj.save(self.filename)

                self.get_info_from_gapi(latitud, longitud, i)

        self.wb_obj.save(self.filename)


folder = pathlib.Path("./excels")
for item in folder.iterdir():
    exc = excelTransformer(f"{item}", sys.argv[1])
    exc.read_excel()
